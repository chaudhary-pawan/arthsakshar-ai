"""
ArthSakshar AI – Excel Tracking Manager
Handles logging call responses to a live Excel tracking sheet.
"""

import os
import threading
from datetime import datetime
import openpyxl

EXCEL_FILE = "call_responses.xlsx"
excel_lock = threading.Lock()

def _init_excel():
    """Initialize the Excel file and headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Responses"
        # Updated Headers based on Prompt
        ws.append(["Name", "Number", "Call Status", "Interest", "Follow Up", "Transferred", "Notes", "Call Time"])
        wb.save(EXCEL_FILE)


def clean_phone(phone: str) -> str:
    """Helper to remove spaces, dashes, or +91 for accurate comparison."""
    if not phone:
        return ""
    p = str(phone).replace(" ", "").replace("-", "")
    if p.startswith("+91"):
        p = p[3:]
    return p[-10:] if len(p) >= 10 else p

def log_call_response(
    phone: str, 
    name: str = "", 
    status: str = "", 
    interest: str = "", 
    follow_up: str = "No", 
    transferred: str = "No", 
    notes: str = ""
):
    """
    Search for a contact by phone number in the Excel sheet and update their row.
    If the contact does not exist, append them as a new row.
    """
    try:
        with excel_lock:
            _init_excel()
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            
            call_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            target_phone_clean = clean_phone(phone)
            
            row_found = False
            
            # Search for the row with this phone number
            # Assuming "Number" is in Column B (index 2)
            for row in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=2).value
                if cell_val and clean_phone(str(cell_val)) == target_phone_clean:
                    # Update the existing row
                    # Headers: Name (1), Number (2), Call Status (3), Interest (4), Follow Up (5), Transferred (6), Notes (7), Call Time (8)
                    if name: ws.cell(row=row, column=1).value = name
                    if status: ws.cell(row=row, column=3).value = status
                    if interest: ws.cell(row=row, column=4).value = interest
                    ws.cell(row=row, column=5).value = follow_up
                    ws.cell(row=row, column=6).value = transferred
                    if notes: ws.cell(row=row, column=7).value = notes
                    ws.cell(row=row, column=8).value = call_time
                    row_found = True
                    break
            
            if not row_found:
                # Append new row if they weren't in the initial spreadsheet
                ws.append([
                    name or "Unknown",
                    phone or "Unknown",
                    status,
                    interest,
                    follow_up,
                    transferred,
                    notes,
                    call_time
                ])
                
            wb.save(EXCEL_FILE)
            
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Failed to log to Excel: {e}")

