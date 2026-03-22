"""
ArthSakshar AI – Bulk Call Manager
Handles batch calling of CA database with rate limiting and status tracking.
"""

import asyncio
import logging
from twilio.rest import Client
import openpyxl
import os
from config import settings
import database as db

EXCEL_FILE = "call_responses.xlsx"

logger = logging.getLogger(__name__)


def get_twilio_client() -> Client:
    """Create Twilio client."""
    return Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)


def load_ca_data(excel_path: str = EXCEL_FILE) -> list[dict]:
    """Load CA database directly from the tracking Excel file."""
    records = []
    try:
        if not os.path.exists(excel_path):
            logger.error(f"Excel file not found at {excel_path}. Please create it and add contacts.")
            return []
            
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Headers are expected: Name (1), Number (2)
        # Start reading from row 1, in case there are no headers
        for row in range(1, ws.max_row + 1):
            name_val = ws.cell(row=row, column=1).value
            phone_val = ws.cell(row=row, column=2).value
            
            # Skip empty rows or header rows
            if not phone_val:
                continue
                
            phone_str = str(phone_val).strip()
            if phone_str.lower() in ("number", "phone", "phone number", "mobile"):
                continue
                
            records.append({
                "name": str(name_val).strip() if name_val else "Unknown",
                "phone": phone_str,
                "city": "", # Can add city column in Excel later if needed
                "language_pref": "marathi", # Defaulting to Marathi per prompt requirement
            })
            
        logger.info(f"Loaded {len(records)} CA records from {excel_path}")
    except Exception as e:
        logger.error(f"Error loading CA data from Excel: {e}")
    return records


def initiate_call(phone_number: str, webhook_url: str) -> str | None:
    """
    Initiate a single outbound call via Twilio.
    Returns call SID or None on failure.
    """
    try:
        client = get_twilio_client()
        call = client.calls.create(
            url=webhook_url,
            to=phone_number,
            from_=settings.TWILIO_PHONE_NUMBER,
            method="POST",
            status_callback=f"{settings.BASE_URL}/voice/status",
            status_callback_method="POST",
            status_callback_event=["initiated", "ringing", "answered", "completed"],
        )
        logger.info(f"Call initiated: {call.sid} → {phone_number}")
        return call.sid
    except Exception as e:
        logger.error(f"Call failed to {phone_number}: {e}")
        return None


async def run_campaign(campaign_id: int, excel_path: str = EXCEL_FILE):
    """
    Run a full bulk calling campaign.
    Batches calls with rate limiting to avoid spam flags.
    """
    records = load_ca_data(excel_path)
    if not records:
        logger.error("No records found in Excel to call")
        await db.update_campaign(campaign_id, status="failed")
        return

    await db.update_campaign(campaign_id, status="running", total_calls=len(records))
    webhook_url = f"{settings.BASE_URL}/voice"

    batch_size = settings.BATCH_SIZE
    delay = settings.BATCH_DELAY_SECONDS
    completed = 0

    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        logger.info(f"Campaign {campaign_id}: Processing batch {i // batch_size + 1} ({len(batch)} calls)")

        for record in batch:
            phone = record["phone"]
            if not phone:
                continue

            call_sid = await asyncio.to_thread(initiate_call, phone, webhook_url)
            if call_sid:
                await db.log_call(
                    call_sid=call_sid,
                    phone_number=phone,
                    ca_name=record.get("name", ""),
                    city=record.get("city", ""),
                    language=record.get("language_pref", "english"),
                )
                completed += 1

            # Small delay between individual calls (2 seconds)
            await asyncio.sleep(2)

        await db.update_campaign(campaign_id, completed_calls=completed)

        # Delay between batches
        if i + batch_size < len(records):
            logger.info(f"Batch complete. Waiting {delay}s before next batch...")
            await asyncio.sleep(delay)

    await db.update_campaign(campaign_id, status="completed", completed_calls=completed)
    logger.info(f"Campaign {campaign_id} completed: {completed}/{len(records)} calls made")
